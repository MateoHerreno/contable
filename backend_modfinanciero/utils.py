from decimal import Decimal
from io import BytesIO
from datetime import datetime
from calendar import monthrange
import openpyxl
import secrets
import pytz

from django.db import transaction
from django.db.models import Sum, OuterRef, Subquery
from django.utils.timezone import make_aware, now
from django.core.mail import send_mail
from django.http import FileResponse
from django.conf import settings

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from .models import (
    Empresa,
    Cliente,
    Proveedor,
    CuentaPorCobrar,
    CuentaPorPagar,
    EstadoResultadosMensual,
    ConceptoCXC,
    ConceptoCXP,
    PermisoPersonalizado,
)
bogota_tz = pytz.timezone('America/Bogota')



#funciones para recuperacion de contraseña por token_______________________________________________________________
def generar_token():
    token = secrets.token_urlsafe(6)  # token ~8 caracteres
    token = token[:8]  # asegúrate de que tenga exactamente 8 
    return f"{token[:4]}-{token[4:]}" #parte el token con un guion

def enviar_email_recuperacion(email, token):
    frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:3000')
    
    asunto = "Recuperación de contraseña"
    mensaje = (
        f"Hola,\n\n"
        f"Has solicitado cambiar tu contraseña.\n"
        f"Tu token de recuperación es:\n\n"
        f"{token}\n\n"
        f"Utiliza este token para cambiar tu contraseña en la aplicación.\n\n"
        f"Tambien puedes ir a la url... \n\n"
        f"{frontend_url}/restablecer?email={email}&token={token}\n\n"
        f"¡Gracias!"
    )
    from_email = None  # usa el DEFAULT_FROM_EMAIL en "contable/contabilidad/contabilidad/setings.py"
    recipient_list = [email]
    send_mail(asunto, mensaje, from_email, recipient_list)

#funcion para devolver con comas los valores (mejor)
def format_decimal_humano(value):
    try:
        return f"{float(value):,.2f}"
    except (ValueError, TypeError):
        return value

#funciones relacionadas con cuentas por cobrar y clientes________________________________________________________
# Actualizar saldos de clientes y proveedores que cargan en cxc y cxp

def recalcular_saldos_proveedor(proveedor_id):
    # Bloquea la fila del proveedor y recalcula en una sola transacción
    with transaction.atomic():
        proveedor = Proveedor.objects.select_for_update().get(pk=proveedor_id)
        saldo_acumulado = Decimal('0')
        cuentas = CuentaPorPagar.objects.filter(
            proveedor_id=proveedor_id
        ).order_by('fecha', 'n_cxp')
        for c in cuentas:
            c.saldo_anterior = saldo_acumulado
            c.pendiente_por_pagar = c.val_bruto + saldo_acumulado - c.abonos
            c.save()
            saldo_acumulado = c.pendiente_por_pagar
        proveedor.saldo = saldo_acumulado
        proveedor.save(update_fields=['saldo'])

def recalcular_saldos_cliente(cliente_id):
    # Bloquea la fila del cliente y recalcula en una sola transacción
    with transaction.atomic():
        cliente = Cliente.objects.select_for_update().get(pk=cliente_id)
        saldo_acumulado = Decimal('0')
        cuentas = CuentaPorCobrar.objects.filter(
            cliente_id=cliente_id
        ).order_by('fecha', 'n_cxc')
        for c in cuentas:
            c.saldo_anterior = saldo_acumulado
            c.pendiente_por_pagar = c.neto_facturado + saldo_acumulado - c.abonos
            c.save()
            saldo_acumulado = c.pendiente_por_pagar
        cliente.saldo = saldo_acumulado
        cliente.save(update_fields=['saldo'])

def recalcular_saldos_todos_clientes():
    # Bloquea todas las filas de clientes antes del bulk_update
    with transaction.atomic():
        sub = CuentaPorCobrar.objects.filter(
            cliente_id=OuterRef('pk')
        ).order_by('-fecha', '-n_cxc')
        qs = Cliente.objects.select_for_update().annotate(
            ultimo_saldo=Subquery(sub.values('pendiente_por_pagar')[:1])
        )
        clientes = list(qs)
        for c in clientes:
            c.saldo = c.ultimo_saldo or Decimal('0')
        Cliente.objects.bulk_update(clientes, ['saldo'])

def recalcular_saldos_todos_proveedores():
    with transaction.atomic():
        sub = CuentaPorPagar.objects.filter(
            proveedor_id=OuterRef('pk')
        ).order_by('-fecha', '-n_cxp')
        qs = Proveedor.objects.select_for_update().annotate(
            ultimo_saldo=Subquery(sub.values('pendiente_por_pagar')[:1])
        )
        proveedores = list(qs)
        for p in proveedores:
            p.saldo = p.ultimo_saldo or Decimal('0')
        Proveedor.objects.bulk_update(proveedores, ['saldo'])

#funcion para exportacion de pdfs______________________________________________________________________________________
#desde cxc filtrando por clientes y fechas (un cliente en un periodo especifico, o si no trae cliente solo la fecha especifica)
def generar_pdf_cxc(queryset, cliente=None):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    ancho_pagina, alto_pagina = letter
    y = alto_pagina - 50

    empresa = Empresa.objects.first()
    nombre_empresa = empresa.nombre if empresa else "Empresa"

    fecha_generacion = now().astimezone(bogota_tz).strftime('%Y-%m-%d %H:%M:%S')
    titulo = f"Reporte Cuentas por Cobrar - {'Cliente: ' + cliente.nombre if cliente else 'General'}"

    pdf.setTitle(titulo)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, nombre_empresa)
    y -= 20
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(50, y, titulo)
    y -= 15
    pdf.setFont("Helvetica", 9)
    pdf.drawString(50, y, f"Generado: {fecha_generacion}")
    y -= 15
    pdf.drawString(50, y, f"Total de cuentas: {queryset.count()}")
    y -= 20

    encabezado = [
        "Fecha", "#CxC", "Cliente", "Concep.", "Bruto",
        "IVA", "Reten.", "Neto Fact.", "Saldo Ant.", "Abonos", "Pend. Pago"
    ]

    x_positions = [40, 90, 140, 250, 320, 370, 420, 470, 520, 570, 620]
    pdf.setFont("Helvetica-Bold", 6.5)
    for i, col in enumerate(encabezado):
        pdf.drawString(x_positions[i], y, col)
    y -= 10
    pdf.line(40, y, 680, y)
    y -= 8

    pdf.setFont("Helvetica", 5)
    total_neto = 0
    total_abonos = 0
    total_pendiente = 0

    for cuenta in queryset:
        datos = [
            cuenta.fecha.strftime('%Y-%m-%d'),
            str(cuenta.n_cxc),
            cuenta.cliente.nombre,
            cuenta.conceptoFijo.nombre,
            f"{cuenta.val_bruto:,.0f}",
            f"{cuenta.iva:,.0f}",
            f"{cuenta.retenciones:,.0f}",
            f"{cuenta.neto_facturado:,.0f}",
            f"{cuenta.saldo_anterior:,.0f}",
            f"{cuenta.abonos:,.0f}",
            f"{cuenta.pendiente_por_pagar:,.0f}"
        ]

        for i, dato in enumerate(datos):
            pdf.drawString(x_positions[i], y, str(dato))

        total_neto += cuenta.neto_facturado
        total_abonos += cuenta.abonos
        total_pendiente += cuenta.pendiente_por_pagar


        y -= 10
        if y < 50:
            pdf.showPage()
            y = alto_pagina - 50

    # Totales finales
    y -= 5
    pdf.line(40, y, 680, y)
    y -= 12
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(40, y, f"Total facturado: ${total_neto:,.2f}")
    y -= 12
    pdf.drawString(40, y, f"Total pagado: ${total_abonos:,.2f}")
    y -= 12
    pdf.drawString(40, y, f"Total adeudado total: ${total_pendiente:,.2f}")

    pdf.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='cuentas_por_cobrar.pdf')

#desde cxp filtrando por fechac y proveedor (si no trae cliente deve traer fecha especifica)
def generar_pdf_cxp(queryset, proveedor=None):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    ancho_pagina, alto_pagina = letter
    y = alto_pagina - 50

    empresa = Empresa.objects.first()
    nombre_empresa = empresa.nombre if empresa else "Empresa"

    fecha_generacion = now().astimezone(bogota_tz).strftime('%Y-%m-%d %H:%M:%S')
    titulo = f"Reporte Cuentas por Pagar "

    pdf.setTitle(titulo)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, nombre_empresa)
    y -= 20
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(50, y, titulo)
    y -= 15
    pdf.setFont("Helvetica", 9)
    pdf.drawString(50, y, f"Generado: {fecha_generacion}")
    y -= 15
    pdf.drawString(50, y, f"Total de cuentas: {queryset.count()}")
    y -= 20

    encabezado = [
        "Fecha", "#CxP", "Proveedor", "Concep.", "Valor",
        "Saldo Ant.", "Abonos", "Pend.Pag."
    ]
    x_positions = [40, 90, 140, 250, 320, 400, 470, 540]

    pdf.setFont("Helvetica-Bold", 6.5)
    for i, col in enumerate(encabezado):
        pdf.drawString(x_positions[i], y, col)
    y -= 10
    pdf.line(40, y, 680, y)
    y -= 8

    pdf.setFont("Helvetica", 5)
    total_bruto = 0
    total_abonos = 0
    total_pendiente = 0

    for cuenta in queryset:
        datos = [
            cuenta.fecha.strftime('%Y-%m-%d'),
            str(cuenta.n_cxp),
            cuenta.proveedor.nombre,
            cuenta.conceptoFijo.nombre,
            f"{cuenta.val_bruto:,.0f}",
            f"{cuenta.saldo_anterior:,.0f}",
            f"{cuenta.abonos:,.0f}",
            f"{cuenta.pendiente_por_pagar:,.0f}"
        ]

        for i, dato in enumerate(datos):
            pdf.drawString(x_positions[i], y, str(dato))

        total_bruto += cuenta.val_bruto
        total_abonos += cuenta.abonos
        total_pendiente += cuenta.pendiente_por_pagar
        y -= 10

        if y < 50:
            pdf.showPage()
            y = alto_pagina - 50

    y -= 5
    pdf.line(40, y, 680, y)
    y -= 12
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(40, y, f"Valor total: ${total_bruto:,.2f}")
    y -= 12
    pdf.drawString(40, y, f"Total abonos: ${total_abonos:,.2f}")
    y -= 12
    pdf.drawString(40, y, f"Total pend. pag. : ${total_pendiente:,.2f}")

    pdf.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='cxp_reporte.pdf')

#exportar pdf estado de resultados
def generar_pdf_estres(anio, mes):
    # Buscar la instancia
    try:
        instancia = EstadoResultadosMensual.objects.get(anio=anio, mes=mes)
    except EstadoResultadosMensual.DoesNotExist:
        from rest_framework.response import Response
        return Response({'error': 'No existe estado de resultados guardado para ese periodo.'}, status=404)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 50

    empresa = Empresa.objects.first()
    nombre_empresa = empresa.nombre if empresa else "Empresa"
    fecha_generacion = now().strftime('%Y-%m-%d %H:%M:%S')
    titulo = f"Estado de Resultados - {mes}/{anio}"

    pdf.setTitle(titulo)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, nombre_empresa)
    y -= 20
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(50, y, titulo)
    y -= 15
    pdf.setFont("Helvetica", 9)
    pdf.drawString(50, y, f"Generado: {fecha_generacion}")
    y -= 25

    def seccion(nombre_seccion, detalle, total, y, mayusculas=True):
        titulo = nombre_seccion.upper() if mayusculas else nombre_seccion
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(50, y, titulo)
        y -= 15
        pdf.setFont("Helvetica", 8)
        for nombre, valor in detalle.items():
            texto = f"{nombre:<40} ${valor:,.2f}"
            pdf.drawRightString(width - 50, y, texto)
            y -= 12
            if y < 50:
                pdf.showPage()
                y = height - 50
        y -= 5
        pdf.setFont("Helvetica", 8.5)
        pdf.drawRightString(width - 50, y, f"total   :  ${total:,.2f}")
        y -= 20
        return y

    y = seccion("INGRESOS", instancia.ingresos_detalle, instancia.ingresos_total, y, mayusculas=True)
    y = seccion("COSTOS DE OPERACIÓN", instancia.costos_operacion_detalle, instancia.costos_operacion_total, y, mayusculas=True)

    # Utilidad bruta alineada a la izquierda en mayúsculas
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawCentredString(width / 2, y, f"UTILIDAD BRUTA: ${instancia.utilidad_bruta:,.2f}")
    y -= 20

    y = seccion("GASTOS ADMINISTRATIVOS", instancia.gastos_administrativos_detalle, instancia.gastos_administrativos_total, y, mayusculas=True)

    # Utilidad operacional alineada a la izquierda en mayúsculas
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawCentredString(width / 2, y, f"UTILIDAD OPERACIONAL: ${instancia.utilidad_operacional:,.2f}")
    y -= 20

    y = seccion("OTROS COSTOS", instancia.otros_costos_detalle, instancia.otros_costos_total, y, mayusculas=True)

    # Utilidad antes de impuestos alineada a la izquierda en mayúsculas
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawCentredString(width / 2, y, f"UTILIDAD ANTES DE IMPUESTOS: ${instancia.utilidad_antes_impuestos:,.2f}")
    y -= 20
    

    y = seccion("GASTOS POR IMPUESTOS", instancia.impuestos_detalle, instancia.gastos_impuestos, y, mayusculas=True)

    # Utilidad neta en mayúsculas y centrada
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawCentredString(width / 2, y, f"UTILIDAD NETA: ${instancia.utilidad_neta:,.2f}")
    y -= 30

    pdf.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'estres_{anio}_{mes}.pdf')


#funciones para exportar a excel______________________________________________________________________________________
def generar_excel_cxc(queryset, cliente=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuentas por Cobrar"

    headers = [
        "Fecha", "#CxC", "Cliente", "Concepto", "Bruto",
        "IVA", "Retenciones", "Neto Facturado", "Saldo Ant.", "Abonos", "Pendiente"
    ]
    ws.append(headers)

    for cuenta in queryset:
        fila = [
            cuenta.fecha.strftime('%Y-%m-%d'),
            cuenta.n_cxc,
            cuenta.cliente.nombre,
            cuenta.conceptoFijo.nombre,
            float(cuenta.val_bruto),
            float(cuenta.iva),
            float(cuenta.retenciones),
            float(cuenta.neto_facturado),
            float(cuenta.saldo_anterior),
            float(cuenta.abonos),
            float(cuenta.pendiente_por_pagar),
        ]
        ws.append(fila)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def generar_excel_cxp(queryset, proveedor=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuentas por Pagar"

    headers = [
        "Fecha", "#CxP", "Proveedor", "Concepto", "Valor",
        "Saldo Ant.", "Abonos", "Pendiente"
    ]
    ws.append(headers)

    for cuenta in queryset:
        fila = [
            cuenta.fecha.strftime('%Y-%m-%d'),
            cuenta.n_cxp,
            cuenta.proveedor.nombre,
            cuenta.conceptoFijo.nombre,
            float(cuenta.val_bruto),
            float(cuenta.saldo_anterior),
            float(cuenta.abonos),
            float(cuenta.pendiente_por_pagar),
        ]
        ws.append(fila)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def generar_excel_estres(instancia):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    ws.append(["Estado de Resultados", f"{instancia.mes}/{instancia.anio}"])
    ws.append(["Fecha de generación", now().strftime('%Y-%m-%d %H:%M:%S')])
    ws.append([])

    def agregar_seccion(titulo, detalle, total):
        ws.append([titulo])
        for key, value in detalle.items():
            ws.append(["", key, float(value)])
        ws.append(["", "Total", float(total)])
        ws.append([])

    agregar_seccion("INGRESOS", instancia.ingresos_detalle, instancia.ingresos_total)
    agregar_seccion("COSTOS DE OPERACIÓN", instancia.costos_operacion_detalle, instancia.costos_operacion_total)
    ws.append(["", "UTILIDAD BRUTA", float(instancia.utilidad_bruta)])
    ws.append([])

    agregar_seccion("GASTOS ADMINISTRATIVOS", instancia.gastos_administrativos_detalle, instancia.gastos_administrativos_total)
    ws.append(["", "UTILIDAD OPERACIONAL", float(instancia.utilidad_operacional)])
    ws.append([])

    agregar_seccion("OTROS COSTOS", instancia.otros_costos_detalle, instancia.otros_costos_total)
    ws.append(["", "UTILIDAD ANTES DE IMPUESTOS", float(instancia.utilidad_antes_impuestos)])
    ws.append([])

    agregar_seccion("GASTOS POR IMPUESTOS", instancia.impuestos_detalle, instancia.gastos_impuestos)
    ws.append(["", "UTILIDAD NETA", float(instancia.utilidad_neta)])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def calcular_y_guardar_estado_resultados(anio, mes):
    fecha_inicio = make_aware(datetime(anio, mes, 1))
    fecha_fin = make_aware(datetime(anio, mes, monthrange(anio, mes)[1], 23, 59, 59))

    cxp_qs = CuentaPorPagar.objects.filter(fecha__range=(fecha_inicio, fecha_fin))
    cxc_qs = CuentaPorCobrar.objects.filter(fecha__range=(fecha_inicio, fecha_fin))

    def sumar_por_concepto(queryset, modelo_concepto):
        resultado = {}
        for concepto in modelo_concepto.objects.all():
            total = queryset.filter(conceptoFijo=concepto).aggregate(suma=Sum('val_bruto'))['suma'] or 0
            resultado[concepto.nombre] = float(total)
        return resultado

    ingresos_detalle = sumar_por_concepto(cxc_qs, ConceptoCXC)
    ingresos_total = sum(ingresos_detalle.values())

    costos_op = ["nomina", "servicios", "otrosOperativos"]
    costos_operacion_detalle = {
        k: float(cxp_qs.filter(conceptoFijo__nombre=k).aggregate(suma=Sum('val_bruto'))['suma'] or 0)
        for k in costos_op
    }
    costos_operacion_total = sum(costos_operacion_detalle.values())
    utilidad_bruta = ingresos_total - costos_operacion_total

    admon = ["honorariosADMON", "honorariosCONT", "segSocial", "otrosADMON"]
    gastos_admon_detalle = {
        k: float(cxp_qs.filter(conceptoFijo__nombre=k).aggregate(suma=Sum('val_bruto'))['suma'] or 0)
        for k in admon
    }
    gastos_admon_total = sum(gastos_admon_detalle.values())
    utilidad_operacional = utilidad_bruta - gastos_admon_total

    otros_costos_detalle = {
        "gastosBancarios": float(cxp_qs.filter(conceptoFijo__nombre="gastosBancarios").aggregate(suma=Sum('val_bruto'))['suma'] or 0)
    }
    otros_costos_total = sum(otros_costos_detalle.values())
    utilidad_antes_impuestos = utilidad_operacional - otros_costos_total

    impuestos_detalle = {
        "impuestos": float(cxc_qs.filter(conceptoFijo__nombre="impuestos").aggregate(suma=Sum('val_bruto'))['suma'] or 0)
    }
    gastos_impuestos = sum(impuestos_detalle.values())
    utilidad_neta = utilidad_antes_impuestos - gastos_impuestos

    estado, _ = EstadoResultadosMensual.objects.update_or_create(
        anio=anio,
        mes=mes,
        defaults={
            "ingresos_total": ingresos_total,
            "ingresos_detalle": ingresos_detalle,
            "costos_operacion_total": costos_operacion_total,
            "costos_operacion_detalle": costos_operacion_detalle,
            "utilidad_bruta": utilidad_bruta,
            "gastos_administrativos_total": gastos_admon_total,
            "gastos_administrativos_detalle": gastos_admon_detalle,
            "utilidad_operacional": utilidad_operacional,
            "otros_costos_total": otros_costos_total,
            "otros_costos_detalle": otros_costos_detalle,
            "utilidad_antes_impuestos": utilidad_antes_impuestos,
            "gastos_impuestos": gastos_impuestos,
            "impuestos_detalle": impuestos_detalle,
            "utilidad_neta": utilidad_neta,
        }
    )
    return estado